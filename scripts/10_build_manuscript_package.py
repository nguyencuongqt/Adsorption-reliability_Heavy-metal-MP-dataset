from __future__ import annotations

import json
import shutil
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image


STUDY_ROOT = Path(__file__).resolve().parents[1]
RESULTS_DIR = STUDY_ROOT / "results"
PACKAGE_DIR = STUDY_ROOT / "manuscript_package"
DATASET_PATH = STUDY_ROOT / "data" / "inputs" / "05_modeling_dataset_final.csv"
CONFIG_PATH = STUDY_ROOT / "configs" / "study_config.json"


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _format_numeric_for_excel(series: pd.Series) -> pd.Series:
    if pd.api.types.is_integer_dtype(series):
        return series
    if pd.api.types.is_numeric_dtype(series):
        return series.round(3)
    return series


def _prettify_table(df: pd.DataFrame) -> pd.DataFrame:
    pretty = df.copy()
    rename_map = {
        "regime": "Validation regime",
        "model": "Model",
        "method": "Importance method",
        "n_splits": "N splits",
        "r2_mean": "Mean R2",
        "r2_sd": "SD R2",
        "rmse_mean": "Mean RMSE",
        "rmse_sd": "SD RMSE",
        "mae_mean": "Mean MAE",
        "mae_sd": "SD MAE",
        "delta_r2_vs_random": "Delta R2 vs random",
        "delta_rmse_vs_random": "Delta RMSE vs random",
        "delta_mae_vs_random": "Delta MAE vs random",
    }
    value_map = {
        "random_cv": "Random CV",
        "group_exp": "Grouped by experiment",
        "group_aut": "Grouped by study",
        "elastic_net": "EN",
        "lightgbm": "LGBM",
        "mlp_regressor": "MLP",
        "elastic_net_coef": "EN coefficient",
        "elastic_net_permutation": "EN permutation",
        "lightgbm_permutation": "LGBM permutation",
        "lightgbm_builtin_gain": "LGBM gain",
        "lightgbm_shap": "LGBM SHAP",
    }
    for column in pretty.columns:
        if pd.api.types.is_string_dtype(pretty[column]) or pretty[column].dtype == object:
            pretty[column] = pretty[column].map(lambda value: value_map.get(value, value))
        pretty[column] = _format_numeric_for_excel(pretty[column])
    return pretty.rename(columns=rename_map)


def style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    sheet = workbook.active
    header_fill = PatternFill(fill_type="solid", fgColor="D9E8FB")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5, header=0.2, footer=0.2)
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 28)
        for cell in column_cells[1:]:
            if isinstance(cell.value, float):
                cell.number_format = "0.000"
            cell.alignment = Alignment(vertical="top", horizontal="center")
    workbook.save(path)


def build_dataset_variable_summary() -> pd.DataFrame:
    df = pd.read_csv(DATASET_PATH)
    variables = [
        ("aut_id", "Study identifier for the source publication", "ID"),
        ("exp_id", "Experiment identifier nested within study", "ID"),
        ("qe", "Equilibrium adsorption capacity of metal on microplastic", "mg g^-1"),
        ("ph", "Solution pH", "unitless"),
        ("temp", "Experimental temperature", "deg C"),
        ("ce", "Equilibrium dissolved metal concentration", "mg L^-1"),
        ("rpm", "Agitation or mixing speed", "rpm"),
        ("sa", "Specific surface area of the microplastic", "m^2 g^-1"),
        ("fg_complexing", "Indicator for complexing surface functional groups", "0/1"),
        ("fg_polar", "Indicator for polar surface functional groups", "0/1"),
        ("fg_any", "Indicator for any annotated surface functional group", "0/1"),
        ("ags_aged", "Indicator for aged microplastic", "0/1"),
        ("ags_virgin", "Indicator for virgin microplastic", "0/1"),
        ("ph_missing", "Indicator that solution pH was not reported", "0/1"),
        ("temp_missing", "Indicator that temperature was not reported", "0/1"),
        ("rpm_missing", "Indicator that mixing speed was not reported", "0/1"),
        ("sa_missing", "Indicator that specific surface area was not reported", "0/1"),
        ("metal_cd", "Indicator for Cd(II)", "0/1"),
        ("metal_cr", "Indicator for Cr(VI)", "0/1"),
        ("metal_hg", "Indicator for Hg(II)", "0/1"),
        ("ret_other", "Indicator for polymer types outside the named categories", "0/1"),
        ("ret_pa", "Indicator for polyamide (PA)", "0/1"),
        ("ret_pe", "Indicator for polyethylene (PE)", "0/1"),
        ("ret_pet", "Indicator for polyethylene terephthalate (PET)", "0/1"),
        ("ret_pla", "Indicator for polylactic acid (PLA)", "0/1"),
        ("ret_pp", "Indicator for polypropylene (PP)", "0/1"),
        ("ret_ps", "Indicator for polystyrene (PS)", "0/1"),
        ("ret_pvc", "Indicator for polyvinyl chloride (PVC)", "0/1"),
    ]
    records = []
    for variable, description, unit in variables:
        missing_pct = float(df[variable].isna().mean() * 100.0) if variable in df.columns else 0.0
        records.append(
            {
                "Variable": variable,
                "Description": description,
                "Unit": unit,
                "Missing %": round(missing_pct, 1),
            }
        )
    return pd.DataFrame(records)


def write_dataset_variable_summary(dst_csv: Path, dst_xlsx: Path) -> None:
    ensure_parent(dst_csv)
    summary = build_dataset_variable_summary()
    summary.to_csv(dst_csv, index=False)
    summary.to_excel(dst_xlsx, index=False)
    style_workbook(dst_xlsx)


def build_augmented_dataset_with_null_features() -> tuple[pd.DataFrame, list[str]]:
    df = pd.read_csv(DATASET_PATH)
    config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    rng = np.random.default_rng(config["null_feature_seed"])
    working = df.copy()

    working["null_random_normal"] = rng.normal(loc=0.0, scale=1.0, size=len(working))
    working["null_random_uniform"] = rng.uniform(low=-1.0, high=1.0, size=len(working))
    working["null_random_binary"] = rng.integers(low=0, high=2, size=len(working))

    for column in ["ph", "ce"]:
        values = working[column].to_numpy(copy=True)
        permuted = values.copy()
        rng.shuffle(permuted)
        working[f"null_perm_{column}"] = permuted

    study_noise = {study: float(rng.normal()) for study in sorted(working["aut_id"].dropna().unique())}
    exp_noise = {exp: float(rng.normal()) for exp in sorted(working["exp_id"].dropna().unique())}
    working["null_study_random_effect"] = working["aut_id"].map(study_noise).astype(float)
    working["null_exp_random_effect"] = working["exp_id"].map(exp_noise).astype(float)

    null_columns = [column for column in working.columns if column.startswith("null_")]
    return working, null_columns


def write_dataframe_table(df: pd.DataFrame, dst_csv: Path, dst_xlsx: Path) -> None:
    ensure_parent(dst_csv)
    df.to_csv(dst_csv, index=False)
    df.to_excel(dst_xlsx, index=False)
    style_workbook(dst_xlsx)


def build_null_feature_catalog() -> pd.DataFrame:
    records = [
        {
            "Name": "null_random_normal",
            "Type": "Pure noise",
            "Construction": "Independent Gaussian noise sampled from N(0,1) for each row.",
        },
        {
            "Name": "null_random_uniform",
            "Type": "Pure noise",
            "Construction": "Independent continuous noise sampled from a uniform distribution on [-1,1] for each row.",
        },
        {
            "Name": "null_random_binary",
            "Type": "Pure noise",
            "Construction": "Independent Bernoulli-like binary noise sampled as 0 or 1 for each row.",
        },
        {
            "Name": "null_perm_ph",
            "Type": "Permuted feature",
            "Construction": "Row-wise permutation of the observed pH values, preserving the marginal distribution while breaking row-level association with qe.",
        },
        {
            "Name": "null_perm_ce",
            "Type": "Permuted feature",
            "Construction": "Row-wise permutation of the observed equilibrium concentration (ce) values, preserving the marginal distribution while breaking row-level association with qe.",
        },
        {
            "Name": "null_study_random_effect",
            "Type": "Group-aware",
            "Construction": "Random normal value assigned at the study level and broadcast to all rows sharing the same aut_id.",
        },
        {
            "Name": "null_exp_random_effect",
            "Type": "Group-aware",
            "Construction": "Random normal value assigned at the experiment level and broadcast to all rows sharing the same exp_id.",
        },
    ]
    return pd.DataFrame(records)


def build_null_feature_summary() -> pd.DataFrame:
    augmented, null_columns = build_augmented_dataset_with_null_features()
    records = []
    for column in null_columns:
        records.append(
            {
                "Name": column,
                "Mean": round(float(augmented[column].mean()), 4),
                "SD": round(float(augmented[column].std(ddof=1)), 4),
                "Pearson corr with qe": round(float(augmented[column].corr(augmented["qe"])), 4),
            }
        )
    return pd.DataFrame(records)


def build_null_feature_example_structure(n_rows: int = 8) -> pd.DataFrame:
    augmented, null_columns = build_augmented_dataset_with_null_features()
    columns = ["aut_id", "exp_id", "qe", "ph", "ce", *null_columns]
    example = augmented.loc[:, columns].head(n_rows).copy()
    example.insert(0, "Example row", range(1, len(example) + 1))
    numeric_columns = example.select_dtypes(include=["number"]).columns
    example.loc[:, numeric_columns] = example.loc[:, numeric_columns].round(4)
    return example


def copy_table(src: Path, dst_csv: Path, dst_xlsx: Path) -> None:
    ensure_parent(dst_csv)
    df = pd.read_csv(src)
    pretty_df = _prettify_table(df)
    pretty_df.to_csv(dst_csv, index=False)
    pretty_df.to_excel(dst_xlsx, index=False)
    style_workbook(dst_xlsx)


def copy_figure_set(src_png: Path, dst_stem: Path) -> None:
    ensure_parent(dst_stem.with_suffix(".png"))
    shutil.copy2(src_png, dst_stem.with_suffix(".png"))
    with Image.open(src_png) as image:
        rgb = image.convert("RGB")
        rgb.save(dst_stem.with_suffix(".pdf"), "PDF", resolution=300.0)
        rgb.save(dst_stem.with_suffix(".tiff"), "TIFF", compression="tiff_lzw")


def main() -> int:
    main_tables = PACKAGE_DIR / "tables"
    si_tables = PACKAGE_DIR / "supplementary_information" / "si_tables"
    main_figures = PACKAGE_DIR / "figures"
    si_figures = PACKAGE_DIR / "supplementary_information" / "si_figures"
    for path in [main_tables, si_tables, main_figures, si_figures]:
        path.mkdir(parents=True, exist_ok=True)

    copy_table(
        RESULTS_DIR / "tables" / "table_dataset_grouped_structure.csv",
        main_tables / "table01_dataset_grouped_structure.csv",
        main_tables / "table01_dataset_grouped_structure.xlsx",
    )
    copy_table(
        RESULTS_DIR / "tables" / "table_performance_across_validation_regimes.csv",
        main_tables / "table02_performance_across_validation_regimes.csv",
        main_tables / "table02_performance_across_validation_regimes.xlsx",
    )
    copy_table(
        RESULTS_DIR / "tables" / "table_generalization_gaps.csv",
        main_tables / "table03_generalization_gaps.csv",
        main_tables / "table03_generalization_gaps.xlsx",
    )
    copy_table(
        RESULTS_DIR / "tables" / "table_feature_importance_method_reliability.csv",
        si_tables / "tableS5_feature_importance_method_reliability.csv",
        si_tables / "tableS5_feature_importance_method_reliability.xlsx",
    )
    copy_table(
        RESULTS_DIR / "tables" / "rq2_mechanism_group_importance.csv",
        si_tables / "tableS6_mechanism_group_importance.csv",
        si_tables / "tableS6_mechanism_group_importance.xlsx",
    )
    copy_table(
        RESULTS_DIR / "tables" / "rq2_feature_importance_summary.csv",
        si_tables / "tableS7_feature_importance_summary.csv",
        si_tables / "tableS7_feature_importance_summary.xlsx",
    )
    write_dataset_variable_summary(
        si_tables / "tableS1_dataset_variable_summary.csv",
        si_tables / "tableS1_dataset_variable_summary.xlsx",
    )
    write_dataframe_table(
        build_null_feature_catalog(),
        si_tables / "tableS2_null_feature_catalog.csv",
        si_tables / "tableS2_null_feature_catalog.xlsx",
    )
    write_dataframe_table(
        build_null_feature_summary(),
        si_tables / "tableS3_null_feature_summary_stats.csv",
        si_tables / "tableS3_null_feature_summary_stats.xlsx",
    )
    write_dataframe_table(
        build_null_feature_example_structure(),
        si_tables / "tableS4_null_feature_example_structure.csv",
        si_tables / "tableS4_null_feature_example_structure.xlsx",
    )

    copy_figure_set(
        RESULTS_DIR / "figures" / "figure_validation_hierarchy_performance_drop.png",
        main_figures / "fig01_validation_hierarchy_performance_drop",
    )
    copy_figure_set(
        RESULTS_DIR / "figures" / "figure_mechanism_group_importance_across_methods.png",
        main_figures / "fig02_mechanism_group_importance_across_methods",
    )
    copy_figure_set(
        RESULTS_DIR / "figures" / "figure_null_feature_benchmarking.png",
        main_figures / "fig03_null_feature_benchmarking",
    )
    copy_figure_set(
        RESULTS_DIR / "figures" / "figure_validation_hierarchy_r2.png",
        si_figures / "sifig01_00_validation_hierarchy_r2",
    )
    copy_figure_set(
        RESULTS_DIR / "figures" / "figure_importance_stability_heatmap.png",
        si_figures / "sifig02_importance_stability_heatmap",
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
